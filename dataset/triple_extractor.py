import os,sys,pathlib
import openpyxl as xl
import json

with open("PartImageNet_Desc.json",'r') as f:
    desc = json.loads(f.read())
f.close()

# triple extract
for obj in desc:
    tri = []
    name = list(obj.keys())[0]
    rel = obj[list(obj.keys())[-1]]
    for key in list(obj[name].keys()):
        for attr in list(obj[name][key].keys())[:-1]:
            c = name + ' has ' + str(obj[name][key][attr]) + ' ' + attr + ' ' + key
            tri.append(c.replace(' Quantity',''))
        tri.append('Especially ' + obj[name][key]['Special Attribute'])
    outpath = os.path.join("triple",name.replace(' ','_').lower()+'.txt')
    with open(outpath,'w') as f:
        for i in tri:
            f.write(i+'\n')
        for r in rel:
            f.write(r+'\n')
    f.close()
    
# excel generate
wb = xl.Workbook()
sheet = wb.active

sheet.append([" Object "," Component "," Description "," Relationship "])
obj_idx = 2
for obj in desc:
    row = []
    name = list(obj.keys())[0]
    rel = str(obj[list(obj.keys())[-1]]).replace(',','\n')
    skip_lines = len(obj[name].keys())-1
    for key in obj[name].keys():
        comp_entry = {}
        comp_entry['comp'] = key
        comp_entry['desc'] = json.dumps(obj[name][key])[1:-1].replace('"'," ")
        row.append(comp_entry)
    this_row = []
    for _row in row:
        this_row = ['',_row['comp'],_row['desc'],'']
        sheet.append(this_row)
    sheet.merge_cells('A{}:A{}'.format(obj_idx,obj_idx+skip_lines))
    # print ('A{}:A{}'.format(obj_idx,obj_idx+skip_lines))
    sheet.merge_cells('D{}:D{}'.format(obj_idx,obj_idx+skip_lines))
    # print ('D{}:D{}'.format(obj_idx,obj_idx+skip_lines))
    sheet['A{}'.format(obj_idx)] = name
    sheet['D{}'.format(obj_idx)] = rel
    obj_idx+=skip_lines+1
    
wb.save("test.xlsx")